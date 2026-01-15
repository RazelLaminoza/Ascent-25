import streamlit as st
import random
import qrcode
import io
import pandas as pd
import base64
import time
import json
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

# ---------------- STORAGE FILE ----------------
DATA_FILE = "raffle_data.json"

# Load previous entries if available
if os.path.exists(DATA_FILE):
    with open(DATA_FILE, "r") as f:
        data = json.load(f)
        st.session_state.entries = data.get("entries", [])
        st.session_state.winner = data.get("winner", None)
else:
    st.session_state.entries = []
    st.session_state.winner = None

if "page" not in st.session_state:
    st.session_state.page = "landing"
if "admin" not in st.session_state:
    st.session_state.admin = False

# ---------------- FUNCTIONS ----------------
def save_data():
    with open(DATA_FILE, "w") as f:
        json.dump({
            "entries": st.session_state.entries,
            "winner": st.session_state.winner
        }, f)

def generate_qr(data):
    qr = qrcode.QRCode(box_size=10, border=4)
    qr.add_data(data)
    qr.make(fit=True)
    return qr.make_image(fill_color="black", back_color="white")

def generate_excel_with_qr(entries, file_path="raffle_entries.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Entries"
    ws.append(["Name", "Employee ID", "QR Code"])
    for idx, e in enumerate(entries, start=2):
        name = e.get("Name", "")
        emp_id = e.get("Employee ID", "")
        ws.cell(row=idx, column=1, value=name)
        ws.cell(row=idx, column=2, value=emp_id)
        qr_data = f"Name: {name}\nEmployee ID: {emp_id}"
        qr_img = qrcode.make(qr_data)
        buf = io.BytesIO()
        qr_img.save(buf, format="PNG")
        buf.seek(0)
        img = XLImage(buf)
        img.width = 100
        img.height = 100
        ws.add_image(img, f"C{idx}")
    wb.save(file_path)

def set_bg_local(image_file):
    with open(image_file, "rb") as f:
        encoded = base64.b64encode(f.read()).decode()
    st.markdown(f"""
    <style>
    [data-testid="stAppViewContainer"] {{
        background-image: url("data:image/png;base64,{encoded}");
        background-size: cover;
        background-position: center;
        background-attachment: fixed;
        font-family: Helvetica, Arial, sans-serif;
    }}
    h1, h2, h3 {{
        color: white;
        text-align: center;
        text-shadow: 1px 1px 2px rgba(0,0,0,0.5);
    }}
    .card {{
        background: rgba(255,255,255,0.2);
        padding: 25px;
        border-radius: 18px;
        backdrop-filter: blur(8px);
        max-width: 400px;
        margin: 20px auto;
        box-shadow: none;
    }}
    .rainbow {{
        animation: rainbow 2s linear infinite;
        font-size: 90px;
        font-weight: bold;
        text-align:center;
    }}
    @keyframes rainbow {{
        0%{{color:#FF0000;}}
        14%{{color:#FF7F00;}}
        28%{{color:#FFFF00;}}
        42%{{color:#00FF00;}}
        57%{{color:#0000FF;}}
        71%{{color:#4B0082;}}
        85%{{color:#8B00FF;}}
        100%{{color:#FF0000;}}
    }}
    div.stTextInput > label {{
        color: white !important;
        font-weight: 600;
    }}
    div.stTextInput > div > input,
    div.stTextInput > div > div > input,
    div.stTextInput > div > textarea {{
        color: black !important;
        background: transparent !important;
        border: none !important;
        border-radius: 0px !important;
        padding: 4px 2px !important;
        font-size: 16px;
        font-family: Helvetica, Arial, sans-serif;
    }}
    div.stTextInput > div > input:focus,
    div.stTextInput > div > div > input:focus,
    div.stTextInput > div > textarea:focus {{
        outline: none;
        border: none;
        background: transparent;
    }}
    </style>
    """, unsafe_allow_html=True)

set_bg_local("bgna.png")

# ---------------- LANDING PAGE ----------------
if st.session_state.page == "landing":
    st.markdown("<h1>Welcome</h1>", unsafe_allow_html=True)
    st.image("welcome_photo.png", use_column_width=True)
    st.markdown("""
    <div class="card">
        <p><strong>Venue:</strong> Okada Manila Ballroom 1–3</p>
        <p><strong>Date:</strong> January 25, 2026</p>
        <p><strong>Time:</strong> 5:00 PM</p>
    </div>
    """, unsafe_allow_html=True)
    if st.button("Register Here"):
        st.session_state.page = "register"

# ---------------- REGISTRATION PAGE ----------------
elif st.session_state.page == "register":
    st.markdown("<h1>Register Here</h1>", unsafe_allow_html=True)
    st.markdown("<div class='card'>", unsafe_allow_html=True)

    # Registration Form
    with st.form("register_form"):
        name = st.text_input("Full Name")
        emp_number = st.text_input("Employee ID")
        submit = st.form_submit_button("Submit")
        if submit:
            if name and emp_number:
                if any(e.get("Employee ID", "") == emp_number for e in st.session_state.entries):
                    st.warning("Employee ID already registered")
                else:
                    # Generate QR
                    qr_img = generate_qr(f"Name: {name}\nEmployee ID: {emp_number}")
                    buf = io.BytesIO()
                    qr_img.save(buf, format="PNG")
                    qr_b64 = base64.b64encode(buf.getvalue()).decode()

                    # Save entry with QR
                    st.session_state.entries.append({
                        "Name": name,
                        "Employee ID": emp_number,
                        "QR": qr_b64
                    })
                    save_data()
                    st.success("Registration successful!")

    st.markdown("</div>", unsafe_allow_html=True)

    # Show current entries with QR immediately
    if st.session_state.entries:
        st.subheader("Current Registered Employees (Editable)")
        df = pd.DataFrame([
            {"Name": e["Name"], "Employee ID": e["Employee ID"]} 
            for e in st.session_state.entries
        ])
        st.data_editor(df, num_rows="dynamic")

        # Display QR images below table
        for e in st.session_state.entries:
            qr_img_html = f'<img src="data:image/png;base64,{e["QR"]}" width="100"/>'
            st.markdown(f"{e['Name']} ({e['Employee ID']})<br>{qr_img_html}", unsafe_allow_html=True)

    if st.button("Admin Login"):
        st.session_state.page = "admin"

# ---------------- ADMIN LOGIN ----------------
elif st.session_state.page == "admin":
    st.markdown("<h2>Admin Login</h2>", unsafe_allow_html=True)
    user = st.text_input("Username")
    pwd = st.text_input("Password", type="password")
    if st.button("⬅ Back to Register"):
        st.session_state.page = "register"
    if st.button("Login"):
        if user == st.secrets["ADMIN_USER"] and pwd == st.secrets["ADMIN_PASS"]:
            st.session_state.admin = True
            st.session_state.page = "raffle"
        else:
            st.error("Invalid credentials")

# ---------------- RAFFLE PAGE ----------------
elif st.session_state.page == "raffle":
    st.markdown("<h1>Raffle Draw</h1>", unsafe_allow_html=True)
    nav1, nav2 = st.columns(2)
    with nav1:
        if st.button("⬅ Register"):
            st.session_state.page = "register"
    with nav2:
        if st.button("🚪 Logout Admin"):
            st.session_state.admin = False
            st.session_state.page = "landing"

    st.subheader("Upload Excel with Name and Employee ID")
    uploaded_file = st.file_uploader("Choose Excel file", type=["xlsx"])
    if uploaded_file:
        df = pd.read_excel(uploaded_file)
        if "Name" in df.columns and "Employee ID" in df.columns:
            st.session_state.entries = [{"Name": r["Name"], "Employee ID": r["Employee ID"]} for _, r in df.iterrows()]
            save_data()
            st.success("Excel uploaded successfully!")
        else:
            st.error("Excel must have 'Name' and 'Employee ID' columns")

    entries = st.session_state.entries
    if entries:
        st.subheader("Registered Employees (Editable)")
        # ---- Editable table ----
        edited_df = st.data_editor(pd.DataFrame(entries), num_rows="dynamic")
        st.session_state.entries = edited_df.to_dict("records")

        # Download Excel with QR
        if st.button("Download Excel with QR"):
            generate_excel_with_qr(st.session_state.entries)
            with open("raffle_entries.xlsx", "rb") as f:
                st.download_button("Download Excel with QR", f, file_name="raffle_entries.xlsx")

        placeholder = st.empty()
        if st.button("Run Raffle"):
            for _ in range(30):
                current = random.choice(st.session_state.entries)
                placeholder.markdown(
                    f"<h1 style='color:white;font-size:70px'>{current.get('Name','')} ({current.get('Employee ID','')})</h1>",
                    unsafe_allow_html=True
                )
                time.sleep(0.07)

            winner = random.choice(st.session_state.entries)
            st.session_state.winner = winner
            save_data()
            placeholder.markdown(
                f"<div class='rainbow'>{winner.get('Name','')} ({winner.get('Employee ID','')})</div>",
                unsafe_allow_html=True
            )

            # Confetti
            confetti_html = """
            <div id="confetti-container"></div>
            <script>
            const colors = ['#FF0000','#FF7F00','#FFFF00','#00FF00','#0000FF','#4B0082','#8B00FF'];
            for(let i=0;i<100;i++){
                let div = document.createElement('div');
                div.className = 'confetti-piece';
                div.style.left = Math.random() * window.innerWidth + 'px';
                div.style.backgroundColor = colors[Math.floor(Math.random()*colors.length)];
                div.style.animationDuration = (Math.random() * 3 + 2) + 's';
                div.style.width = div.style.height = (Math.random() * 8 + 4) + 'px';
                document.body.appendChild(div);
            }
            setTimeout(()=>{document.getElementById('confetti-container').remove()}, 4000);
            </script>
            """
            st.components.v1.html(confetti_html, height=0, width=0)
    else:
        st.info("No entries yet. Upload an Excel file first.")
